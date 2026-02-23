#!/usr/bin/env python3
"""
Prepare normalized datasets from source Excel files.

Outputs:
  - output/spreadsheet/cases_dataset_clean.xlsx
  - output/spreadsheet/cases_import.csv
  - output/spreadsheet/import_report.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def norm_space(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.strip()
    if text.lower() in {"nan", "nat", "none", "null"}:
        return ""
    return text


def norm_for_key(value: Any) -> str:
    return norm_space(value).lower()


def parse_decimal(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = norm_space(value)
    if text == "":
        return None

    text = text.replace(" ", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", ".", "-", "-."}:
        return None

    try:
        return round(float(text), 2)
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.strftime("%Y-%m-%d")

    text = norm_space(value)
    if text == "":
        return None

    text = text.replace("г.", "").replace("г", "").strip()
    text = text.replace("/", ".")

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            dt = pd.to_datetime(text, format=fmt, errors="raise")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass

    # Date ranges and textual windows: keep start date.
    match = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
    if match:
        try:
            dt = pd.to_datetime(match.group(1), format="%d.%m.%Y", errors="raise")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None

    return None


def parse_contract_date(contract_ref: Any) -> str | None:
    text = norm_space(contract_ref)
    if text == "":
        return None

    match = re.search(r"\bот\s+(\d{2}\.\d{2}\.\d{2,4})\b", text, flags=re.IGNORECASE)
    if not match:
        return None
    raw = match.group(1)
    if len(raw.rsplit(".", 1)[-1]) == 2:
        try:
            dt = pd.to_datetime(raw, format="%d.%m.%y", errors="raise")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None
    try:
        dt = pd.to_datetime(raw, format="%d.%m.%Y", errors="raise")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def parse_year_from_text(text: str) -> int | None:
    match = re.search(r"\b(20\d{2})\b", text)
    if not match:
        return None
    return int(match.group(1))


def parse_case_code(subject: str) -> str | None:
    if subject == "":
        return None

    patterns = [
        r"\b\d{2}[/-]\d{2}[/-][A-Za-zА-Яа-яЁё]{1,5}-\d+\b",
        r"\b\d{2}[/-]\d{2}-[A-Za-zА-Яа-яЁё]{1,5}-\d+\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, subject)
        if match:
            return match.group(0)

    # fallback: first token if it looks like code with many separators
    first = subject.split(" ", 1)[0]
    if first.count("-") >= 2 and re.search(r"\d", first):
        return first

    return None


def infer_result_status(result_raw: Any) -> str | None:
    text = norm_for_key(result_raw)
    if text == "":
        return None

    if "неисполн" in text or "нов" in text or "не начат" in text:
        return "NEW"
    if "отмен" in text or "снят" in text:
        return "CANCELLED"
    if "в работе" in text or "частич" in text or "передано" in text or "исполня" in text:
        return "IN_PROGRESS"
    if "исполн" in text or "выполн" in text or "закрыт" in text:
        return "DONE"
    if "без исполн" in text:
        return "NO_ACTION"
    return None


def normalize_assignees(raw: Any) -> tuple[str | None, list[str]]:
    text = norm_space(raw)
    if text == "":
        return None, []

    # canonical mapping for common typos/variants
    aliases = {
        "сенотова": "Сенотова",
        "шатина": "Шатина",
        "сликина": "Слинкина",
        "мнедведев": "Медведев",
        "медведева": "Медведев",
        "мозер/шашкин": "Мозер; Шашкин",
        "медведев / мозер": "Медведев; Мозер",
        "шашкин слинкина": "Шашкин; Слинкина",
        "слинкина шашкин": "Слинкина; Шашкин",
        "медведев шатина": "Медведев; Шатина",
        "слинкина и макарова": "Слинкина; Макарова",
    }

    key = norm_for_key(text)
    text = aliases.get(key, text)

    text = re.sub(r"\s*(/|;|,|\sи\s)\s*", ";", text, flags=re.IGNORECASE)
    parts = [norm_space(x) for x in text.split(";")]
    parts = [x for x in parts if x]

    # Title-case fallback for Latin/Cyrillic text chunks
    normalized = []
    for part in parts:
        lower = norm_for_key(part)
        mapped = aliases.get(lower, part)
        normalized.append(mapped)

    if not normalized:
        return None, []
    return "; ".join(normalized), normalized


def block_type_by_sheet(sheet_name: str) -> str:
    key = norm_for_key(sheet_name)
    if "расторжен" in key:
        return "TERMINATIONS"
    if "претенз" in key or "притенз" in key:
        return "CLAIMS"
    if "заключ" in key:
        return "CONCLUDED"
    if "утвержден" in key:
        return "APPROVED_FZ"
    return "TASKS"


def col_by_contains(columns: list[str], needle: str) -> str | None:
    n = norm_for_key(needle)
    for col in columns:
        if n in norm_for_key(col):
            return col
    return None


def choose_col(columns: list[str], *candidates: str) -> str | None:
    for candidate in candidates:
        col = col_by_contains(columns, candidate)
        if col:
            return col
    return None


def safe_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return int(value)
    text = norm_space(value)
    if text == "":
        return None
    m = re.search(r"-?\d+", text)
    if not m:
        return None
    return int(m.group(0))


def make_bundle_key(prefix: str, values: list[Any]) -> str:
    seed = "|".join(norm_for_key(v) for v in values if norm_space(v) != "")
    if seed == "":
        seed = "empty"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:20]
    return f"{prefix}|{digest}"


@dataclass
class NormalizeResult:
    cases_rows: list[dict[str, Any]]
    incoming_rows: list[dict[str, Any]]
    claims_rows: list[dict[str, Any]]
    issues: list[dict[str, Any]]


def normalize_tasks_workbook(path: Path) -> NormalizeResult:
    xls = pd.ExcelFile(path)
    cases_rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    file_year = parse_year_from_text(path.name)

    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        df.columns = [norm_space(c) for c in df.columns]
        columns = list(df.columns)

        task_col = choose_col(columns, "ЗАДАЧА")
        if not task_col:
            continue

        reg_col = choose_col(columns, "№ п/п2", "№ п/п")
        article_col = choose_col(columns, "Статья")
        form_col = choose_col(columns, "форма закупки")
        amount_col = choose_col(columns, "Сумма")
        rnmc_col = choose_col(columns, "Сумма РНМЦК")
        task_date_col = choose_col(columns, "Дата поставленной задачи")
        due_col = choose_col(columns, "Планируемая дата исполнения")
        notes_col = choose_col(columns, "Примечания", "Примечание")
        archive_col = choose_col(columns, "ФЗ")
        assignee_col = choose_col(columns, "ФИО исполнителя")
        status_col = choose_col(columns, "Исполнено/неисполнено", "Исполнено/ неисполнено")
        contract_col = choose_col(columns, "№ контракта")
        contract_sum_col = choose_col(columns, "Сумма контракта")
        rts_col = choose_col(columns, "№ РТС", "№РТС")
        procurement_no_col = choose_col(columns, "№ закупки")

        for idx, row in df.iterrows():
            subject_raw = norm_space(row.get(task_col))
            if subject_raw == "":
                continue

            reg_no = safe_int(row.get(reg_col)) if reg_col else None
            task_date = parse_date(row.get(task_date_col)) if task_date_col else None
            due_date = parse_date(row.get(due_col)) if due_col else None

            assignees_norm, assignee_parts = normalize_assignees(row.get(assignee_col) if assignee_col else None)
            status_raw = norm_space(row.get(status_col)) if status_col else ""
            status_raw = status_raw or None
            status = infer_result_status(status_raw)

            contract_number = norm_space(row.get(contract_col)) if contract_col else ""
            contract_number = contract_number or None
            contract_date = parse_contract_date(contract_number)

            archive_path = norm_space(row.get(archive_col)) if archive_col else ""
            archive_path = archive_path or None

            case_code = parse_case_code(subject_raw)
            year = parse_year_from_text(sheet)
            if year is None and task_date:
                year = int(task_date[:4])
            if year is None and contract_date:
                year = int(contract_date[:4])
            if year is None and file_year:
                year = file_year

            if year is None:
                issues.append(
                    {
                        "source": "tasks",
                        "sheet": sheet,
                        "row": int(idx + 2),
                        "type": "missing_year",
                        "subject": subject_raw[:120],
                    }
                )

            notes_parts = []
            notes_val = norm_space(row.get(notes_col)) if notes_col else ""
            if notes_val:
                notes_parts.append(notes_val)
            rts = norm_space(row.get(rts_col)) if rts_col else ""
            if rts:
                notes_parts.append(f"РТС: {rts}")
            pr_no = norm_space(row.get(procurement_no_col)) if procurement_no_col else ""
            if pr_no:
                notes_parts.append(f"Закупка: {pr_no}")

            bundle_key = make_bundle_key(
                "tasks",
                [
                    block_type_by_sheet(sheet),
                    case_code,
                    subject_raw,
                    contract_number,
                    archive_path,
                ],
            )

            cases_rows.append(
                {
                    "source_file": path.name,
                    "source_sheet": sheet,
                    "source_row": int(idx + 2),
                    "source_kind": "tasks_registry",
                    "block_type": block_type_by_sheet(sheet),
                    "year": year,
                    "reg_no": reg_no,
                    "case_code": case_code,
                    "subject_raw": subject_raw,
                    "subject_clean": None,
                    "budget_article": norm_space(row.get(article_col)) or None if article_col else None,
                    "procurement_form": norm_space(row.get(form_col)) or None if form_col else None,
                    "amount_planned": parse_decimal(row.get(amount_col)) if amount_col else None,
                    "rnmc_amount": parse_decimal(row.get(rnmc_col)) if rnmc_col else None,
                    "task_date": task_date,
                    "stage_raw": None,
                    "due_date": due_date,
                    "notes": " | ".join(notes_parts) if notes_parts else None,
                    "archive_path": archive_path,
                    "result_raw": status_raw,
                    "result_status": status,
                    "result_amount": None,
                    "result_percent": None,
                    "contract_ref_raw": contract_number,
                    "contract_number": contract_number,
                    "contract_date": contract_date,
                    "contract_amount": parse_decimal(row.get(contract_sum_col)) if contract_sum_col else None,
                    "bundle_key": bundle_key,
                    "assignees_text": assignees_norm,
                    "assignees_list": "; ".join(assignee_parts) if assignee_parts else None,
                }
            )

    return NormalizeResult(cases_rows=cases_rows, incoming_rows=[], claims_rows=[], issues=issues)


def normalize_claims_registry(path: Path) -> NormalizeResult:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    year: int | None = None
    reg_no = 0
    for r in range(1, ws.max_row + 1):
        contractor = ws.cell(r, 1).value
        contract_ref = ws.cell(r, 2).value
        claimed = ws.cell(r, 3).value
        paid = ws.cell(r, 4).value
        payment_ref = ws.cell(r, 5).value
        writeoff_ref = ws.cell(r, 6).value
        note = ws.cell(r, 7).value

        first = norm_space(contractor)
        if re.fullmatch(r"20\d{2}\s*год", first):
            year = int(first[:4])
            reg_no = 0
            continue

        if first and norm_space(contract_ref) and parse_decimal(claimed) is not None:
            reg_no += 1
            claim_val = parse_decimal(claimed)
            paid_val = parse_decimal(paid)
            contract_text = norm_space(contract_ref)
            contract_date = parse_contract_date(contract_text)

            status_raw_parts = []
            if paid_val is not None:
                status_raw_parts.append(f"Оплачено {paid_val}")
            if norm_space(writeoff_ref):
                status_raw_parts.append("Списание по ПП 783")
            if not status_raw_parts:
                status_raw_parts.append("Претензия выставлена")
            result_raw = "; ".join(status_raw_parts)

            notes_parts = []
            if norm_space(payment_ref):
                notes_parts.append(f"Платежка: {norm_space(payment_ref)}")
            if norm_space(writeoff_ref):
                notes_parts.append(norm_space(writeoff_ref))
            if norm_space(note):
                notes_parts.append(norm_space(note))

            bundle_key = make_bundle_key(
                "claims_registry",
                [year, contractor, contract_ref, claimed],
            )

            rows.append(
                {
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": r,
                    "source_kind": "claims_registry",
                    "block_type": "CLAIMS",
                    "year": year,
                    "reg_no": reg_no,
                    "case_code": None,
                    "subject_raw": f"Претензия к {first} по контракту {contract_text}",
                    "subject_clean": None,
                    "budget_article": None,
                    "procurement_form": None,
                    "amount_planned": claim_val,
                    "rnmc_amount": None,
                    "task_date": None,
                    "stage_raw": None,
                    "due_date": None,
                    "notes": " | ".join(notes_parts) if notes_parts else None,
                    "archive_path": None,
                    "result_raw": result_raw,
                    "result_status": "DONE" if (paid_val is not None or norm_space(writeoff_ref)) else None,
                    "result_amount": paid_val if paid_val is not None else None,
                    "result_percent": 100 if (paid_val is not None or norm_space(writeoff_ref)) else None,
                    "contract_ref_raw": contract_text,
                    "contract_number": contract_text,
                    "contract_date": contract_date,
                    "contract_amount": None,
                    "bundle_key": bundle_key,
                    "assignees_text": None,
                    "assignees_list": None,
                }
            )
        elif first and norm_space(contract_ref) == "" and parse_decimal(claimed) is None:
            # likely a header/extra line; ignore silently
            continue

    # quality checks
    if not rows:
        issues.append({"source": "claims_registry", "type": "no_rows_parsed", "file": path.name})

    return NormalizeResult(cases_rows=rows, incoming_rows=[], claims_rows=rows, issues=issues)


def normalize_incoming_registry(path: Path) -> NormalizeResult:
    xls = pd.ExcelFile(path)
    rows: list[dict[str, Any]] = []
    cases_rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        df.columns = [norm_space(c) for c in df.columns]
        columns = list(df.columns)

        in_no_col = choose_col(columns, "№ входящего")
        in_date_col = choose_col(columns, "Дата входящего")
        title_col = choose_col(columns, "Краткое наименование")
        rez_col = choose_col(columns, "Резолюция")
        exec_col = choose_col(columns, "Исполнитель")

        if not title_col:
            continue

        for idx, row in df.iterrows():
            title = norm_space(row.get(title_col))
            if title == "":
                continue

            in_no = norm_space(row.get(in_no_col)) if in_no_col else ""
            in_date = parse_date(row.get(in_date_col)) if in_date_col else None
            rez_date = parse_date(row.get(rez_col)) if rez_col else None
            assignees_norm, _ = normalize_assignees(row.get(exec_col) if exec_col else None)

            sheet_year = parse_year_from_text(sheet)
            row_year = int(in_date[:4]) if in_date else sheet_year

            if sheet_year and in_date and int(in_date[:4]) != sheet_year:
                issues.append(
                    {
                        "source": "incoming_registry",
                        "sheet": sheet,
                        "row": int(idx + 2),
                        "type": "sheet_year_mismatch",
                        "incoming_date": in_date,
                    }
                )

            rows.append(
                {
                    "source_file": path.name,
                    "source_sheet": sheet,
                    "source_row": int(idx + 2),
                    "year": row_year,
                    "incoming_no": in_no or None,
                    "incoming_date": in_date,
                    "title": title,
                    "resolution_date": rez_date,
                    "assignees_text": assignees_norm,
                }
            )

            case_year = row_year
            incoming_no_num = safe_int(in_no)
            case_code = None
            if case_year is not None and incoming_no_num is not None:
                case_code = f"IN-{case_year}-{incoming_no_num}"
            elif incoming_no_num is not None:
                case_code = f"IN-{incoming_no_num}"

            notes_parts = [f"Входящий №{in_no}" if in_no else "Входящий документ"]
            if rez_date:
                notes_parts.append(f"Резолюция: {rez_date}")
            notes = " | ".join(notes_parts)

            bundle_key = make_bundle_key(
                "incoming_registry",
                [sheet, in_no, in_date, title],
            )

            cases_rows.append(
                {
                    "source_file": path.name,
                    "source_sheet": sheet,
                    "source_row": int(idx + 2),
                    "source_kind": "incoming_registry",
                    "block_type": "TASKS",
                    "year": case_year,
                    "reg_no": incoming_no_num,
                    "case_code": case_code,
                    "subject_raw": title,
                    "subject_clean": None,
                    "budget_article": None,
                    "procurement_form": "INCOMING_DOC",
                    "amount_planned": None,
                    "rnmc_amount": None,
                    "task_date": in_date,
                    "stage_raw": "INCOMING_REGISTRY",
                    "due_date": rez_date,
                    "notes": notes,
                    "archive_path": None,
                    "result_raw": f"Резолюция {rez_date}" if rez_date else "Входящий документ",
                    "result_status": "IN_PROGRESS" if rez_date else "NEW",
                    "result_amount": None,
                    "result_percent": None,
                    "contract_ref_raw": None,
                    "contract_number": None,
                    "contract_date": None,
                    "contract_amount": None,
                    "bundle_key": bundle_key,
                    "assignees_text": assignees_norm,
                    "assignees_list": assignees_norm,
                }
            )

    return NormalizeResult(cases_rows=cases_rows, incoming_rows=rows, claims_rows=[], issues=issues)


def write_outputs(
    out_dir: Path,
    cases_rows: list[dict[str, Any]],
    incoming_rows: list[dict[str, Any]],
    claims_rows: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)

    cases_df = pd.DataFrame(cases_rows)
    incoming_df = pd.DataFrame(incoming_rows)
    claims_df = pd.DataFrame(claims_rows)
    issues_df = pd.DataFrame(issues)

    # import CSV for the PHP importer
    import_cols = [
        "source_file",
        "source_sheet",
        "source_row",
        "source_kind",
        "block_type",
        "year",
        "reg_no",
        "case_code",
        "subject_raw",
        "subject_clean",
        "budget_article",
        "procurement_form",
        "amount_planned",
        "rnmc_amount",
        "task_date",
        "stage_raw",
        "due_date",
        "notes",
        "archive_path",
        "result_raw",
        "result_status",
        "result_amount",
        "result_percent",
        "contract_ref_raw",
        "contract_number",
        "contract_date",
        "contract_amount",
        "bundle_key",
        "assignees_text",
    ]

    for col in import_cols:
        if col not in cases_df.columns:
            cases_df[col] = None
    import_df = cases_df[import_cols].copy()

    csv_path = out_dir / "cases_import.csv"
    import_df.to_csv(csv_path, index=False, encoding="utf-8")

    xlsx_path = out_dir / "cases_dataset_clean.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        cases_df.to_excel(writer, sheet_name="cases_normalized", index=False)
        incoming_df.to_excel(writer, sheet_name="incoming_docs_normalized", index=False)
        claims_df.to_excel(writer, sheet_name="claims_registry_normalized", index=False)
        issues_df.to_excel(writer, sheet_name="quality_issues", index=False)

    duplicate_bundle_count = 0
    if not import_df.empty:
        duplicate_bundle_count = int(import_df["bundle_key"].duplicated().sum())

    report = {
        "cases_rows_total": int(len(cases_df)),
        "cases_rows_by_block_type": {
            str(k): int(v) for k, v in cases_df["block_type"].value_counts(dropna=False).to_dict().items()
        }
        if not cases_df.empty
        else {},
        "incoming_docs_rows_total": int(len(incoming_df)),
        "claims_registry_rows_total": int(len(claims_df)),
        "issues_total": int(len(issues_df)),
        "duplicate_bundle_key_rows": duplicate_bundle_count,
        "paths": {
            "xlsx": str(xlsx_path),
            "csv": str(csv_path),
        },
    }

    report_path = out_dir / "import_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare normalized cases dataset from source Excel files")
    parser.add_argument("--incoming", required=True, help="Path to incoming docs registry xlsx")
    parser.add_argument("--claims", required=True, help="Path to claims registry xlsx")
    parser.add_argument("--tasks", required=True, help="Path to tasks registry xlsx")
    parser.add_argument(
        "--out-dir",
        default="output/spreadsheet",
        help="Output directory (default: output/spreadsheet)",
    )
    args = parser.parse_args()

    incoming_path = Path(args.incoming).expanduser().resolve()
    claims_path = Path(args.claims).expanduser().resolve()
    tasks_path = Path(args.tasks).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()

    if not incoming_path.exists():
        raise FileNotFoundError(f"Incoming registry not found: {incoming_path}")
    if not claims_path.exists():
        raise FileNotFoundError(f"Claims registry not found: {claims_path}")
    if not tasks_path.exists():
        raise FileNotFoundError(f"Tasks registry not found: {tasks_path}")

    tasks_result = normalize_tasks_workbook(tasks_path)
    claims_result = normalize_claims_registry(claims_path)
    incoming_result = normalize_incoming_registry(incoming_path)

    cases_rows = tasks_result.cases_rows + claims_result.cases_rows + incoming_result.cases_rows
    incoming_rows = incoming_result.incoming_rows
    claims_rows = claims_result.claims_rows
    issues = tasks_result.issues + claims_result.issues + incoming_result.issues

    report = write_outputs(out_dir, cases_rows, incoming_rows, claims_rows, issues)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
